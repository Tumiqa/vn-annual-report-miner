# -*- coding: utf-8 -*-
"""
arminer.export.financial_excel
==============================
Tao file Excel bao cao tai chinh chuyen nghiep:
- Tab Bao_Cao_Tai_Chinh: Trinh bay day du toan bo 700+ chi tieu, phan chia ro rang theo 13 nhom chuan muc:
    1. CĐKT. TÀI SẢN NGẮN HẠN
    2. CĐKT. TÀI SẢN DÀI HẠN
    3. CĐKT. NỢ PHẢI TRẢ NGẮN HẠN
    4. CĐKT. NỢ PHẢI TRẢ DÀI HẠN
    5. CĐKT. VỐN CHỦ SỞ HỮU
    6. KQKD. DOANH THU, CHI PHÍ, LỢI NHUẬN
    7. LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG KINH DOANH
    8. LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG ĐẦU TƯ
    9. LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG TÀI CHÍNH
    10. LCTT. DÒNG TIỀN THUẦN, TIỀN CUỐI KÌ
    11. NGOẠI BẢNG. A TÀI SẢN CỦA CTCK VÀ TÀI SẢN QUẢN LÝ THEO CAM KẾT
    12. NGOẠI BẢNG. B TÀI SẢN VÀ CÁC KHOẢN PHẢI TRẢ VỀ TÀI SẢN QUẢN LÝ CAM KẾT VỚI KHÁCH HÀNG
    13. THUYẾT MINH. CÁC LOẠI TÀI SẢN TÀI CHÍNH
- Tab Ty_So_Tai_Chinh: He thong chi so tai chinh toan dien chuan WiData (Kha nang sinh loi, Don bay & Thanh toan,
  Dac thu CTCK / Margin / Dau tu, Toc do tang truong YoY %, Quy mo & Dinh gia).
- Tab Panel_Data_Goc: Bang du lieu bang phang (Panel Data) chuan nghien cuu kinh te luong.
- Tab Codebook: Tu dien bien chi tiet.
- Tab Huong_Dan_VBA: Huong dan su dung bo loc va ma nguon VBA.
Xuat ca file .xlsx (chuan) va file .xlsm (tich hop Macro VBA va cac nut bam loc nhanh).
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from loguru import logger

# 13 Nhom chuan hoa bao cao tai chinh theo dung yeu cau
CATEGORY_ORDER = [
    "CĐKT. TÀI SẢN NGẮN HẠN",
    "CĐKT. TÀI SẢN DÀI HẠN",
    "CĐKT. NỢ PHẢI TRẢ NGẮN HẠN",
    "CĐKT. NỢ PHẢI TRẢ DÀI HẠN",
    "CĐKT. VỐN CHỦ SỞ HỮU",
    "KQKD. DOANH THU, CHI PHÍ, LỢI NHUẬN",
    "LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG KINH DOANH",
    "LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG ĐẦU TƯ",
    "LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG TÀI CHÍNH",
    "LCTT. DÒNG TIỀN THUẦN, TIỀN CUỐI KÌ",
    "NGOẠI BẢNG. A TÀI SẢN CỦA CTCK VÀ TÀI SẢN QUẢN LÝ THEO CAM KẾT",
    "NGOẠI BẢNG. B TÀI SẢN VÀ CÁC KHOẢN PHẢI TRẢ VỀ TÀI SẢN QUẢN LÝ CAM KẾT VỚI KHÁCH HÀNG",
    "THUYẾT MINH. CÁC LOẠI TÀI SẢN TÀI CHÍNH",
]
CATEGORIES = CATEGORY_ORDER

def classify_financial_item(code: str, name: str, stmt: str, order: int = 0) -> str:
    """Phan loai chi tieu vao dung 1 trong 13 nhom theo quy chuan ke toan Viet Nam."""
    c = str(code).lower()
    n = str(name).lower()

    # 1. Thuyet minh
    if "thuyet_minh" in c or "thuyết minh" in n:
        return "THUYẾT MINH. CÁC LOẠI TÀI SẢN TÀI CHÍNH"

    # 2. Ngoai bang
    if "ngoai_bang" in c or "ngoại bảng" in n:
        if any(k in c or k in n for k in ["khach_hang", "nha_dau_tu", "khách hàng", "nhà đầu tư", "phải trả", "phai_tra"]):
            return "NGOẠI BẢNG. B TÀI SẢN VÀ CÁC KHOẢN PHẢI TRẢ VỀ TÀI SẢN QUẢN LÝ CAM KẾT VỚI KHÁCH HÀNG"
        return "NGOẠI BẢNG. A TÀI SẢN CỦA CTCK VÀ TÀI SẢN QUẢN LÝ THEO CAM KẾT"

    # 3. Ket qua kinh doanh
    if stmt == "income_statement":
        return "KQKD. DOANH THU, CHI PHÍ, LỢI NHUẬN"

    # 4. Luu chuyen tien te
    if stmt == "cash_flow":
        if any(k in c or k in n for k in ["đầu tư", "dau_tu", "mua_sam", "thanh_ly", "cho_vay", "tien_gui", "thu_lai"]):
            return "LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG ĐẦU TƯ"
        elif any(k in c or k in n for k in ["tài chính", "tai_chinh", "co_tuc", "cổ tức", "von_gop", "vốn góp", "vay", "tra_no", "cổ phiếu quỹ"]):
            return "LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG TÀI CHÍNH"
        elif any(k in c or k in n for k in ["thuần trong kỳ", "thuan_trong_ky", "đầu kỳ", "dau_ky", "cuối kỳ", "cuoi_ky", "tỷ giá", "ty_gia", "tiền cuối kỳ"]):
            return "LCTT. DÒNG TIỀN THUẦN, TIỀN CUỐI KÌ"
        else:
            return "LCTT. DÒNG TIỀN TỪ HOẠT ĐỘNG KINH DOANH"

    # 5. Bang can doi ke toan
    if stmt == "balance_sheet":
        if any(k in c or k in n for k in ["vốn chủ sở hữu", "von_chu_so_huu", "vốn đầu tư của chủ sở hữu", "thặng dư", "cổ phiếu quỹ", "nguồn kinh phí", "lợi nhuận sau thuế chưa phân phối"]):
            return "CĐKT. VỐN CHỦ SỞ HỮU"

        if any(k in c or k in n for k in ["nợ dài hạn", "no_dai_han", "vay và nợ thuê tài chính dài hạn", "trái phiếu phát hành dài hạn"]):
            return "CĐKT. NỢ PHẢI TRẢ DÀI HẠN"

        if any(k in c or k in n for k in ["nợ ngắn hạn", "no_ngan_han", "vay và nợ thuê tài chính ngắn hạn", "trái phiếu phát hành ngắn hạn", "chi phí phải trả ngắn hạn", "phải trả người bán ngắn hạn", "người mua trả tiền trước ngắn hạn"]):
            return "CĐKT. NỢ PHẢI TRẢ NGẮN HẠN"

        if any(k in c or k in n for k in ["tài sản dài hạn", "tai_san_dai_han", "tài sản cố định", "bất động sản đầu tư", "xây dựng cơ bản"]):
            return "CĐKT. TÀI SẢN DÀI HẠN"

        if any(k in c or k in n for k in ["tài sản ngắn hạn", "tai_san_ngan_han", "tiền và tương đương", "tiền và các khoản", "đầu tư ngắn hạn", "chứng khoán kinh doanh", "phải thu ngắn hạn", "hàng tồn kho"]):
            return "CĐKT. TÀI SẢN NGẮN HẠN"

        if "nợ" in n or "phải trả" in n or "no_" in c:
            if "dài hạn" in n or "dai_han" in c:
                return "CĐKT. NỢ PHẢI TRẢ DÀI HẠN"
            return "CĐKT. NỢ PHẢI TRẢ NGẮN HẠN"

        if "dài hạn" in n or "dai_han" in c or order > 120:
            return "CĐKT. TÀI SẢN DÀI HẠN"

        return "CĐKT. TÀI SẢN NGẮN HẠN"

    return "CĐKT. TÀI SẢN NGẮN HẠN"


# He thong dinh nghia toan bo cac chi so WiData
WIDATA_RATIOS = {
    # 1. Kha nang sinh loi & Hieu qua
    "roa": {
        "name": "ROA (%) (Y)",
        "group": "Khả năng sinh lời",
        "formula": "Lợi nhuận sau thuế / Tổng tài sản",
        "fmt": "0.00%",
    },
    "roe": {
        "name": "ROE (%) (Y)",
        "group": "Khả năng sinh lời",
        "formula": "Lợi nhuận sau thuế / Vốn chủ sở hữu",
        "fmt": "0.00%",
    },
    "gross_margin": {
        "name": "Biên lợi nhuận gộp (%) (Y)",
        "group": "Khả năng sinh lời",
        "formula": "Lợi nhuận gộp / Doanh thu thuần",
        "fmt": "0.00%",
    },
    "net_margin": {
        "name": "Biên lợi nhuận ròng (%) (Y)",
        "group": "Khả năng sinh lời",
        "formula": "Lợi nhuận sau thuế / Doanh thu thuần",
        "fmt": "0.00%",
    },
    "ebit_margin": {
        "name": "Biên EBIT (%) (Y)",
        "group": "Khả năng sinh lời",
        "formula": "EBIT / Doanh thu thuần",
        "fmt": "0.00%",
    },
    "effective_tax_rate": {
        "name": "Tỷ lệ thuế suất hiệu dụng (%) (Y)",
        "group": "Khả năng sinh lời",
        "formula": "Chi phí thuế TNDN / Lợi nhuận trước thuế",
        "fmt": "0.00%",
    },
    "asset_turnover": {
        "name": "Vòng quay tổng tài sản (Lần) (Y)",
        "group": "Hiệu quả hoạt động",
        "formula": "Doanh thu thuần / Tổng tài sản",
        "fmt": "0.00",
    },
    "cfo_to_net_income": {
        "name": "Dòng tiền HĐKD/Lợi nhuận thuần (%) (Y)",
        "group": "Hiệu quả dòng tiền",
        "formula": "Dòng tiền thuần HĐKD / Lợi nhuận sau thuế",
        "fmt": "0.00%",
    },
    "cfo_to_avg_assets": {
        "name": "Dòng tiền HĐKD/Trung bình tổng tài sản (%) (Y)",
        "group": "Hiệu quả dòng tiền",
        "formula": "Dòng tiền thuần HĐKD / Tổng tài sản",
        "fmt": "0.00%",
    },
    "cfo_to_avg_equity": {
        "name": "Dòng tiền HĐKD/Trung bình vốn chủ sỡ hữu (%) (Y)",
        "group": "Hiệu quả dòng tiền",
        "formula": "Dòng tiền thuần HĐKD / Vốn chủ sở hữu",
        "fmt": "0.00%",
    },

    # 2. Co cau von & Kha nang thanh toan
    "debt_to_assets": {
        "name": "Hệ số nợ trên tổng tài sản (%) (Y)",
        "group": "Cơ cấu vốn và Đòn bẩy",
        "formula": "Nợ phải trả / Tổng tài sản",
        "fmt": "0.00%",
    },
    "debt_to_equity": {
        "name": "Hệ số nợ trên vốn chủ sở hữu (%) (Y)",
        "group": "Cơ cấu vốn và Đòn bẩy",
        "formula": "Nợ phải trả / Vốn chủ sở hữu",
        "fmt": "0.00%",
    },
    "equity_to_assets": {
        "name": "Hệ số vốn chủ sở hữu (%) (Y)",
        "group": "Cơ cấu vốn và Đòn bẩy",
        "formula": "Vốn chủ sở hữu / Tổng tài sản",
        "fmt": "0.00%",
    },
    "equity_multiplier": {
        "name": "Tổng tài sản/Vốn chủ sở hữu (Lần) (Y)",
        "group": "Cơ cấu vốn và Đòn bẩy",
        "formula": "Tổng tài sản / Vốn chủ sở hữu",
        "fmt": "0.00",
    },
    "current_ratio": {
        "name": "Tỷ số thanh toán hiện hành (Lần) (Y)",
        "group": "Khả năng thanh toán",
        "formula": "Tài sản ngắn hạn / Nợ ngắn hạn",
        "fmt": "0.00",
    },
    "quick_ratio": {
        "name": "Tỷ số thanh toán nhanh (Lần) (Y)",
        "group": "Khả năng thanh toán",
        "formula": "(Tài sản ngắn hạn - Hàng tồn kho) / Nợ ngắn hạn",
        "fmt": "0.00",
    },

    # 3. Dac thu CTCK & Tai san tai chinh
    "margin_to_equity": {
        "name": "Tỷ lệ cho vay ký quỹ trên VCSH (%) (Y)",
        "group": "Đặc thù CTCK & Margin",
        "formula": "Cho vay ký quỹ (margin) / Vốn chủ sở hữu",
        "fmt": "0.00%",
    },
    "pct_margin_loans": {
        "name": "% Cho vay nghiệp vụ ký quỹ (margin) (%) (Y)",
        "group": "Đặc thù CTCK & Margin",
        "formula": "Cho vay margin / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_advances": {
        "name": "% Cho vay ứng trước tiền bán chứng khoán của khách hàng (%) (Y)",
        "group": "Đặc thù CTCK & Margin",
        "formula": "Cho vay ứng trước tiền bán CK / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_fvtpl": {
        "name": "% Tài sản tài chính ghi nhận thông qua lãi lỗ (FVTPL) (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Tài sản FVTPL / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_afs": {
        "name": "% Tài sản tài chính sẵn sàng để bán (AFS) (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Tài sản AFS / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_htm": {
        "name": "% Tài sản tài chính giữ đến ngày đáo hạn (HTM) (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Tài sản HTM / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_cash": {
        "name": "% Tiền và các khoản tương đương tiền (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Tiền và tương đương tiền / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_loans": {
        "name": "% Các khoản cho vay (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Tổng các khoản cho vay / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_brokerage_rev": {
        "name": "% Doanh thu hoạt động môi giới chứng khoán (%) (Y)",
        "group": "Cơ cấu doanh thu",
        "formula": "Doanh thu môi giới CK / Tổng doanh thu hoạt động",
        "fmt": "0.00%",
    },
    "pct_proprietary_rev": {
        "name": "% Doanh thu mảng tự doanh và kinh doanh nguồn vốn (%) (Y)",
        "group": "Cơ cấu doanh thu",
        "formula": "Doanh thu tự doanh / Tổng doanh thu hoạt động",
        "fmt": "0.00%",
    },
    "pct_margin_profit": {
        "name": "% Lợi nhuận cho vay ký quỹ (%) (Y)",
        "group": "Cơ cấu lợi nhuận",
        "formula": "Lợi nhuận cho vay margin / Lợi nhuận hoạt động",
        "fmt": "0.00%",
    },
    "pct_ib_rev": {
        "name": "% Doanh thu mảng ngân hàng đầu tư (%) (Y)",
        "group": "Cơ cấu doanh thu",
        "formula": "Doanh thu tư vấn tài chính, IB / Tổng doanh thu",
        "fmt": "0.00%",
    },
    "pct_brokerage_cost": {
        "name": "% Chi phí hoạt động môi giới chứng khoán (%) (Y)",
        "group": "Cơ cấu chi phí",
        "formula": "Chi phí môi giới / Tổng chi phí hoạt động",
        "fmt": "0.00%",
    },
    "pct_proprietary_cost": {
        "name": "% Chi phí hoạt động tự doanh (%) (Y)",
        "group": "Cơ cấu chi phí",
        "formula": "Chi phí tự doanh / Tổng chi phí hoạt động",
        "fmt": "0.00%",
    },
    "pct_advisory_cost": {
        "name": "% Chi phí hoạt động tư vấn tài chính (%) (Y)",
        "group": "Cơ cấu chi phí",
        "formula": "Chi phí tư vấn tài chính / Tổng chi phí hoạt động",
        "fmt": "0.00%",
    },
    "pct_provision_cost": {
        "name": "% Chi phí dự phòng/Hoàn nhập TSTC (%) (Y)",
        "group": "Cơ cấu chi phí",
        "formula": "Chi phí dự phòng TSTC / Tổng chi phí hoạt động",
        "fmt": "0.00%",
    },

    # 4. Tang truong cung ky YoY (%)
    "rev_growth_yoy": {
        "name": "Doanh thu (*) (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Doanh thu T - Doanh thu T-1) / |Doanh thu T-1|",
        "fmt": "0.00%",
    },
    "ebt_growth_yoy": {
        "name": "Lợi nhuận trước thuế (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(LNTT T - LNTT T-1) / |LNTT T-1|",
        "fmt": "0.00%",
    },
    "eat_growth_yoy": {
        "name": "Lợi nhuận sau thuế (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(LNST T - LNST T-1) / |LNST T-1|",
        "fmt": "0.00%",
    },
    "eat_parent_growth_yoy": {
        "name": "Lợi nhuận sau thuế CĐCTM (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(LNST Cty mẹ T - LNST Cty mẹ T-1) / |LNST Cty mẹ T-1|",
        "fmt": "0.00%",
    },
    "assets_growth_yoy": {
        "name": "Tổng tài sản (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Tổng tài sản T - Tổng tài sản T-1) / Tổng tài sản T-1",
        "fmt": "0.00%",
    },
    "equity_growth_yoy": {
        "name": "Vốn chủ sở hữu (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Vốn CSH T - Vốn CSH T-1) / Vốn CSH T-1",
        "fmt": "0.00%",
    },
    "debt_growth_yoy": {
        "name": "Nợ phải trả (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Nợ phải trả T - Nợ phải trả T-1) / Nợ phải trả T-1",
        "fmt": "0.00%",
    },
    "margin_loans_growth_yoy": {
        "name": "Cho vay nghiệp vụ ký quỹ (margin) (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Dư nợ margin T - Dư nợ margin T-1) / Dư nợ margin T-1",
        "fmt": "0.00%",
    },
    "advances_growth_yoy": {
        "name": "Cho vay ứng trước tiền bán chứng khoán của khách hàng (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Ứng trước T - Ứng trước T-1) / Ứng trước T-1",
        "fmt": "0.00%",
    },
    "brokerage_rev_growth_yoy": {
        "name": "Doanh thu hoạt động môi giới chứng khoán (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(DT môi giới T - DT môi giới T-1) / DT môi giới T-1",
        "fmt": "0.00%",
    },
    "proprietary_rev_growth_yoy": {
        "name": "Doanh thu mảng tự doanh và kinh doanh nguồn vốn (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(DT tự doanh T - DT tự doanh T-1) / DT tự doanh T-1",
        "fmt": "0.00%",
    },
    "cash_growth_yoy": {
        "name": "Tiền và các khoản tương đương tiền (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Tiền T - Tiền T-1) / Tiền T-1",
        "fmt": "0.00%",
    },
    "fvtpl_growth_yoy": {
        "name": "Tài sản tài chính ghi nhận thông qua lãi lỗ (FVTPL) (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(FVTPL T - FVTPL T-1) / FVTPL T-1",
        "fmt": "0.00%",
    },
    "htm_growth_yoy": {
        "name": "Tài sản tài chính giữ đến ngày đáo hạn (HTM) (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(HTM T - HTM T-1) / HTM T-1",
        "fmt": "0.00%",
    },
    "afs_growth_yoy": {
        "name": "Tài sản tài chính sẵn sàng để bán (AFS) (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(AFS T - AFS T-1) / AFS T-1",
        "fmt": "0.00%",
    },

    # 5. Quy mo & Dinh gia co ban (VND & ln)
    "total_assets": {
        "name": "Tổng tài sản (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Tổng tài sản",
        "fmt": "#,##0",
    },
    "total_debt": {
        "name": "Nợ phải trả (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Nợ phải trả",
        "fmt": "#,##0",
    },
    "equity": {
        "name": "Vốn chủ sở hữu (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Vốn chủ sở hữu",
        "fmt": "#,##0",
    },
    "net_revenue": {
        "name": "Doanh thu (*) (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Doanh thu thuần / Doanh thu hoạt động",
        "fmt": "#,##0",
    },
    "profit_before_tax": {
        "name": "Lợi nhuận trước thuế (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Lợi nhuận trước thuế",
        "fmt": "#,##0",
    },
    "profit_after_tax": {
        "name": "Lợi nhuận sau thuế (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Lợi nhuận sau thuế",
        "fmt": "#,##0",
    },
    "operating_cash_flow": {
        "name": "Dòng tiền từ hoạt động kinh doanh (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "LCTT - Dòng tiền thuần từ HĐKD",
        "fmt": "#,##0",
    },
    "investing_cash_flow": {
        "name": "Dòng tiền từ hoạt động đầu tư (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "LCTT - Dòng tiền thuần từ HĐĐT",
        "fmt": "#,##0",
    },
    "financing_cash_flow": {
        "name": "Dòng tiền từ hoạt động tài chính (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "LCTT - Dòng tiền thuần từ HĐTC",
        "fmt": "#,##0",
    },
    "eat_parent": {
        "name": "Lợi nhuận sau thuế công ty mẹ (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - LNST Cổ đông công ty mẹ",
        "fmt": "#,##0",
    },
    "curr_debt": {
        "name": "Nợ phải trả ngắn hạn (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Nợ ngắn hạn",
        "fmt": "#,##0",
    },
    "long_term_debt": {
        "name": "Nợ phải trả dài hạn (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Nợ dài hạn",
        "fmt": "#,##0",
    },
    "curr_assets": {
        "name": "Tài sản ngắn hạn (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Tài sản ngắn hạn",
        "fmt": "#,##0",
    },
    "non_curr_assets": {
        "name": "Tài sản dài hạn (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "BCTC - Tài sản dài hạn",
        "fmt": "#,##0",
    },
    "brokerage_profit": {
        "name": "Lợi nhuận từ nghiệp vụ môi giới (VND) (Y)",
        "group": "Đặc thù CTCK & Margin",
        "formula": "Doanh thu môi giới - Chi phí môi giới",
        "fmt": "#,##0",
    },
    "advisory_profit": {
        "name": "Lợi nhuận từ nghiệp vụ tư vấn tài chính (VND) (Y)",
        "group": "Đặc thù CTCK & Margin",
        "formula": "Doanh thu tư vấn - Chi phí tư vấn",
        "fmt": "#,##0",
    },
    "margin_profit": {
        "name": "Lợi nhuận từ cho vay ký quỹ (VND) (Y)",
        "group": "Đặc thù CTCK & Margin",
        "formula": "Lãi từ các khoản cho vay và phải thu",
        "fmt": "#,##0",
    },
    "operating_profit": {
        "name": "Lợi nhuận hoạt động (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "Doanh thu hoạt động - Chi phí hoạt động",
        "fmt": "#,##0",
    },
    "operating_cost": {
        "name": "Chi phí hoạt động (*) (VND) (Y)",
        "group": "Quy mô tài chính",
        "formula": "Tổng chi phí hoạt động CTCK",
        "fmt": "#,##0",
    },
    "curr_debt_growth_yoy": {
        "name": "Nợ phải trả ngắn hạn (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Nợ NH T - Nợ NH T-1) / Nợ NH T-1",
        "fmt": "0.00%",
    },
    "long_debt_growth_yoy": {
        "name": "Nợ phải trả dài hạn (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Nợ DH T - Nợ DH T-1) / Nợ DH T-1",
        "fmt": "0.00%",
    },
    "curr_assets_growth_yoy": {
        "name": "Tài sản ngắn hạn (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(TSNH T - TSNH T-1) / TSNH T-1",
        "fmt": "0.00%",
    },
    "non_curr_assets_growth_yoy": {
        "name": "Tài sản dài hạn (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(TSDH T - TSDH T-1) / TSDH T-1",
        "fmt": "0.00%",
    },
    "oper_cost_growth_yoy": {
        "name": "Chi phí hoạt động (*) (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Chi phí HĐ T - Chi phí HĐ T-1) / Chi phí HĐ T-1",
        "fmt": "0.00%",
    },
    "oper_profit_growth_yoy": {
        "name": "Lợi nhuận hoạt động (YoY) (%) (Y)",
        "group": "Tăng trưởng cùng kỳ (YoY)",
        "formula": "(Lợi nhuận HĐ T - Lợi nhuận HĐ T-1) / Lợi nhuận HĐ T-1",
        "fmt": "0.00%",
    },
    "pct_other_receivables": {
        "name": "% Phải thu khác (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Phải thu khác / Tổng tài sản",
        "fmt": "0.00%",
    },
    "pct_broker_services": {
        "name": "% Phải thu các dịch vụ CTCK cung cấp (%) (Y)",
        "group": "Cơ cấu tài sản tài chính",
        "formula": "Phải thu dịch vụ CTCK / Tổng tài sản",
        "fmt": "0.00%",
    },
    "size_ln": {
        "name": "Quy mô doanh nghiệp Size ln(Tổng tài sản)",
        "group": "Quy mô tài chính",
        "formula": "ln(Tổng tài sản)",
        "fmt": "0.000",
    },
}


def compute_widata_metrics(pivot: pd.DataFrame) -> pd.DataFrame:
    """Tinh toan toan bo he thong chi so WiData tu bang pivot panel."""
    df = pivot.copy()
    if "ticker" not in df.columns or "year" not in df.columns:
        return df

    df = df.sort_values(["ticker", "year"]).reset_index(drop=True)

    def find_col(patterns):
        for p in patterns:
            for c in df.columns:
                if c.lower() == p.lower():
                    return df[c]
        for p in patterns:
            for c in df.columns:
                if p.lower() in c.lower():
                    return df[c]
        return pd.Series(np.nan, index=df.index)

    # Core base items
    total_assets = find_col(["bs_tong_tai_san", "bs_tong_cong_tai_san"])
    equity = find_col(["bs_von_chu_so_huu", "bs_von_chu_so_huu_4d280b22", "bs_von_chu_so_huu_6cda78ae"])
    debt = find_col(["bs_no_phai_tra", "bs_tong_no_phai_tra"])
    curr_assets = find_col(["bs_tai_san_ngan_han"])
    curr_liab = find_col(["bs_no_ngan_han"])
    inventory = find_col(["bs_hang_ton_kho"])
    cash = find_col(["bs_tien_va_tuong_duong_tien", "bs_tien"])
    revenue = find_col(["is_doanh_thu_thuan", "is_doanh_so_thuan", "is_doanh_thu_hoat_dong"])
    gross_profit = find_col(["is_loi_nhuan_gop", "is_lai_gop"])
    ebt = find_col(["is_loi_nhuan_truoc_thue", "is_lai_truoc_thue", "is_tong_loi_nhuan_ke_toan_truoc_thue"])
    eat = find_col(["is_loi_nhuan_sau_thue", "is_lai_sau_thue", "is_loi_nhuan_sau_thue_thu_nhap_doanh_nghiep"])
    eat_parent = find_col(["is_loi_nhuan_sau_thue_cua_co_dong_cong_ty_me", "is_loi_nhuan_sau_thue_cty_me"])
    if eat_parent.isna().all():
        eat_parent = eat
    ebit = find_col(["is_ebit"])
    tax = find_col(["is_chi_phi_thue_tndn", "is_chi_phi_thue_thu_nhap_doanh_nghiep_hien_hanh"])
    cfo = find_col(["cf_luu_chuyen_tien_thuan_tu_cac_hoat_dong_san_xuat_kinh_doanh", "cf_luu_chuyen_tien_thuan_tu_hoat_dong_kinh_doanh"])
    cfi = find_col(["cf_luu_chuyen_tien_thuan_tu_hoat_dong_dau_tu"])
    cff = find_col(["cf_luu_chuyen_tien_thuan_tu_hoat_dong_tai_chinh"])

    # CTCK items
    margin_loans = find_col(["bs_cac_khoan_cho_vay", "bs_phai_thu_ve_cho_vay_ky_quy"])
    advances = find_col(["bs_phai_thu_ung_truoc_tien_ban_chung_khoan_cua_khach_hang", "bs_ung_truoc_tien_ban"])
    fvtpl = find_col(["bs_cac_tai_san_tai_chinh_ghi_nhan_thong_qua_lai_lo_fvtpl"])
    htm = find_col(["bs_cac_khoan_dau_tu_nam_giu_den_ngay_dao_han_htm", "bs_dau_tu_nam_giu_den_ngay_dao_han_htm"])
    afs = find_col(["bs_cac_khoan_tai_chinh_san_sang_de_ban_afs", "bs_tai_san_tai_chinh_san_sang_de_ban_afs"])
    brokerage_rev = find_col(["is_doanh_thu_hoat_dong_moi_gioi_chung_khoan"])
    proprietary_rev = find_col(["is_doanh_thu_mang_tu_doanh_va_kinh_doanh_nguon_von", "is_lai_tu_cac_tai_san_tai_chinh_ghi_nhan_thong_qua_lai_lo_fvtpl"])
    margin_profit = find_col(["is_lai_tu_cac_khoan_cho_vay_va_phai_thu"])
    ib_rev = find_col(["is_doanh_thu_hoat_dong_tu_van_tai_chinh", "is_doanh_thu_mang_ngan_hang_dau_tu"])
    brokerage_cost = find_col(["is_chi_phi_hoat_dong_moi_gioi_chung_khoan"])
    proprietary_cost = find_col(["is_chi_phi_hoat_dong_tu_doanh"])
    advisory_cost = find_col(["is_chi_phi_hoat_dong_tu_van_tai_chinh"])
    provision_cost = find_col(["is_chi_phi_du_phong_tstc", "is_chi_phi_du_phong"])
    long_term_debt = find_col(["bs_no_dai_han", "bs_tong_no_dai_han"])
    non_curr_assets = find_col(["bs_tai_san_dai_han"])
    operating_cost = find_col(["is_chi_phi_hoat_dong", "is_tong_chi_phi_hoat_dong"])
    other_receivables = find_col(["bs_phai_thu_khac", "bs_cac_khoan_phai_thu_khac"])
    broker_services = find_col(["bs_phai_thu_cac_dich_vu_ctck_cung_cap"])

    def sdiv(a, b):
        return a.astype(float) / b.replace(0, np.nan).astype(float)

    # 1. Sinh loi
    df["roa"] = sdiv(eat, total_assets)
    df["roe"] = sdiv(eat, equity)
    df["gross_margin"] = sdiv(gross_profit, revenue)
    df["net_margin"] = sdiv(eat, revenue)
    df["ebit_margin"] = sdiv(ebit, revenue)
    df["effective_tax_rate"] = sdiv(tax, ebt)
    df["asset_turnover"] = sdiv(revenue, total_assets)
    df["cfo_to_net_income"] = sdiv(cfo, eat)
    df["cfo_to_avg_assets"] = sdiv(cfo, total_assets)
    df["cfo_to_avg_equity"] = sdiv(cfo, equity)

    # 2. Don bay & Thanh toan
    df["debt_to_assets"] = sdiv(debt, total_assets)
    df["debt_to_equity"] = sdiv(debt, equity)
    df["equity_to_assets"] = sdiv(equity, total_assets)
    df["equity_multiplier"] = sdiv(total_assets, equity)
    df["current_ratio"] = sdiv(curr_assets, curr_liab)
    df["quick_ratio"] = sdiv(curr_assets - inventory.fillna(0), curr_liab)

    # 3. CTCK & Co cau
    df["margin_to_equity"] = sdiv(margin_loans, equity)
    df["pct_margin_loans"] = sdiv(margin_loans, total_assets)
    df["pct_advances"] = sdiv(advances, total_assets)
    df["pct_fvtpl"] = sdiv(fvtpl, total_assets)
    df["pct_afs"] = sdiv(afs, total_assets)
    df["pct_htm"] = sdiv(htm, total_assets)
    df["pct_cash"] = sdiv(cash, total_assets)
    df["pct_loans"] = sdiv(margin_loans, total_assets)
    df["pct_brokerage_rev"] = sdiv(brokerage_rev, revenue)
    df["pct_proprietary_rev"] = sdiv(proprietary_rev, revenue)
    df["pct_margin_profit"] = sdiv(margin_profit, ebt)
    df["pct_ib_rev"] = sdiv(ib_rev, revenue)
    df["pct_brokerage_cost"] = sdiv(brokerage_cost, revenue)
    df["pct_proprietary_cost"] = sdiv(proprietary_cost, revenue)
    df["pct_advisory_cost"] = sdiv(advisory_cost, revenue)
    df["pct_provision_cost"] = sdiv(provision_cost, revenue)
    df["pct_other_receivables"] = sdiv(other_receivables, total_assets)
    df["pct_broker_services"] = sdiv(broker_services, total_assets)

    # 4. Quy mo
    df["total_assets"] = total_assets
    df["total_debt"] = debt
    df["equity"] = equity
    df["curr_assets"] = curr_assets
    df["non_curr_assets"] = non_curr_assets
    df["curr_debt"] = curr_liab
    df["long_term_debt"] = long_term_debt
    df["net_revenue"] = revenue
    df["profit_before_tax"] = ebt
    df["profit_after_tax"] = eat
    df["eat_parent"] = eat_parent
    df["brokerage_profit"] = brokerage_rev.fillna(0) - brokerage_cost.fillna(0)
    df["advisory_profit"] = ib_rev.fillna(0) - advisory_cost.fillna(0)
    df["margin_profit"] = margin_profit
    df["operating_cost"] = operating_cost
    df["operating_profit"] = revenue.fillna(0) - operating_cost.fillna(0)
    df["operating_cash_flow"] = cfo
    df["investing_cash_flow"] = cfi
    df["financing_cash_flow"] = cff
    df["size_ln"] = total_assets.apply(lambda x: math.log(x) if pd.notna(x) and x > 0 else np.nan)

    # 5. YoY Growth %
    for col, yoy_col in [
        ("net_revenue", "rev_growth_yoy"),
        ("profit_before_tax", "ebt_growth_yoy"),
        ("profit_after_tax", "eat_growth_yoy"),
        ("eat_parent", "eat_parent_growth_yoy"),
        ("total_assets", "assets_growth_yoy"),
        ("equity", "equity_growth_yoy"),
        ("total_debt", "debt_growth_yoy"),
        ("curr_debt", "curr_debt_growth_yoy"),
        ("long_term_debt", "long_debt_growth_yoy"),
        ("curr_assets", "curr_assets_growth_yoy"),
        ("non_curr_assets", "non_curr_assets_growth_yoy"),
        ("operating_cost", "oper_cost_growth_yoy"),
        ("operating_profit", "oper_profit_growth_yoy"),
    ]:
        try:
            df[yoy_col] = df.groupby("ticker")[col].pct_change(fill_method=None)
        except TypeError:
            df[yoy_col] = df.groupby("ticker")[col].pct_change()

    # CTCK specific YoY
    try:
        if margin_loans.name:
            df["margin_loans_growth_yoy"] = df.groupby("ticker")[margin_loans.name].pct_change(fill_method=None)
        if cash.name:
            df["cash_growth_yoy"] = df.groupby("ticker")[cash.name].pct_change(fill_method=None)
        if fvtpl.name:
            df["fvtpl_growth_yoy"] = df.groupby("ticker")[fvtpl.name].pct_change(fill_method=None)
        if htm.name:
            df["htm_growth_yoy"] = df.groupby("ticker")[htm.name].pct_change(fill_method=None)
        if afs.name:
            df["afs_growth_yoy"] = df.groupby("ticker")[afs.name].pct_change(fill_method=None)
    except Exception:
        pass

    return df


def populate_financial_sheets(
    wb: openpyxl.Workbook,
    all_data: pd.DataFrame,
    pivot: pd.DataFrame,
    ratio_cols: Dict[str, str],
    fin_codebook: List[Dict[str, Any]],
) -> openpyxl.Workbook:
    """Xay dung va dien du lieu vao cac sheet BCTC theo dung 13 nhom va ty so WiData."""
    # Tu dong lam giau toan bo he thong ty so WiData
    pivot = compute_widata_metrics(pivot)

    NAVY_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    TEAL_FILL = PatternFill(start_color="205375", end_color="205375", fill_type="solid")
    WHITE_BOLD = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    SECTION_FONT = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    REG_FONT = Font(name="Segoe UI", size=10)
    BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
    ITALIC_FONT = Font(name="Segoe UI", size=9, italic=True, color="555555")

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    tickers = sorted([str(t) for t in pivot["ticker"].dropna().unique().tolist()])
    years = sorted([int(y) for y in pivot["year"].dropna().unique().tolist()])

    ticker_list_str = '"TẤT CẢ,' + ",".join(tickers) + '"'
    dv = DataValidation(type="list", formula1=ticker_list_str, allow_blank=True)

    # -----------------------------------------------------------------
    # Sheet 1: Bao_Cao_Tai_Chinh
    # -----------------------------------------------------------------
    if "Bao_Cao_Tai_Chinh" in wb.sheetnames:
        ws_bc = wb["Bao_Cao_Tai_Chinh"]
        if ws_bc.views.sheetView:
            ws_bc.views.sheetView[0].showGridLines = True
    else:
        ws_bc = wb.create_sheet("Bao_Cao_Tai_Chinh")

    ws_bc["A1"] = "BÁO CÁO TÀI CHÍNH DOANH NGHIỆP"
    ws_bc["A1"].font = TITLE_FONT

    ws_bc["A2"] = "Lọc theo Mã CK:"
    ws_bc["A2"].font = BOLD_FONT
    ws_bc["A2"].alignment = Alignment(horizontal="right")

    ws_bc["B2"] = "TẤT CẢ"
    ws_bc["B2"].font = BOLD_FONT
    ws_bc["B2"].alignment = Alignment(horizontal="center")
    ws_bc["B2"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws_bc.add_data_validation(dv)
    dv.add(ws_bc["B2"])

    # Tieu de cot tai dong 4
    headers_bc = ["Mã CK", "Phân nhóm báo cáo", "Mã chỉ tiêu", "Tên chỉ tiêu"] + [str(y) for y in years]
    for col_idx, h in enumerate(headers_bc, 1):
        cell = ws_bc.cell(row=4, column=col_idx, value=h)
        cell.font = WHITE_BOLD
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # -----------------------------------------------------------------
    # Đảm bảo xuất đủ 100% toàn bộ 702 chỉ tiêu cho mỗi mã chứng khoán
    # -----------------------------------------------------------------
    try:
        import vnfinancialdata as vnf
        df_master = vnf.list_items(active_only=False).copy()
    except Exception as e:
        logger.warning(f"Không thể tải từ điển vnfinancialdata: {e}")
        df_master = pd.DataFrame()

    if df_master.empty:
        if not all_data.empty:
            df_master = all_data[["item_code", "item_name", "statement"]].drop_duplicates().copy()
            df_master["item_order"] = 0
        else:
            df_master = pd.DataFrame(columns=["item_code", "item_name", "statement", "item_order"])

    if "statement" not in df_master.columns:
        df_master["statement"] = df_master["item_code"].apply(
            lambda c: "balance_sheet" if str(c).startswith("bs_") else "income_statement" if str(c).startswith("is_") else "cash_flow"
        )
    if "item_order" not in df_master.columns:
        df_master["item_order"] = 0

    df_master["category"] = df_master.apply(
        lambda r: classify_financial_item(r["item_code"], r["item_name"], r["statement"], r.get("item_order", 0)),
        axis=1
    )

    cat_order_map = {cat: idx for idx, cat in enumerate(CATEGORY_ORDER)}
    df_master["cat_order"] = df_master["category"].map(lambda c: cat_order_map.get(c, 99))
    df_master = df_master.sort_values(["cat_order", "item_order", "item_code"]).reset_index(drop=True)

    # Nếu người dùng chọn riêng một tập hợp chỉ tiêu, lọc chính xác theo tập hợp đó
    if fin_codebook:
        selected_items = {item.get("Biến") for item in fin_codebook if item.get("Phân loại") == "Chỉ tiêu kế toán"}
        if selected_items and len(selected_items) < len(df_master):
            df_master = df_master[df_master["item_code"].isin(selected_items)].copy()

    # Bảng tra cứu số liệu thực tế theo (ticker, item_code, year) -> value
    data_lookup = {}
    if not all_data.empty:
        for _, r in all_data.iterrows():
            t_key = str(r["ticker"]).strip().upper()
            icode_key = str(r["item_code"]).strip()
            try:
                y_key = int(r["year"])
            except (ValueError, TypeError):
                continue
            v_val = r.get("value")
            if pd.notna(v_val) and v_val is not None:
                data_lookup[(t_key, icode_key, y_key)] = v_val

    row_curr = 5
    for t in tickers:
        for _, mrow in df_master.iterrows():
            icode = str(mrow["item_code"])
            ws_bc.cell(row=row_curr, column=1, value=t).alignment = Alignment(horizontal="center")
            ws_bc.cell(row=row_curr, column=2, value=str(mrow["category"])).alignment = Alignment(horizontal="left")
            ws_bc.cell(row=row_curr, column=3, value=icode).alignment = Alignment(horizontal="left")
            ws_bc.cell(row=row_curr, column=4, value=str(mrow["item_name"])).alignment = Alignment(horizontal="left")

            for y_idx, y in enumerate(years, 5):
                val = data_lookup.get((t, icode, y))
                c = ws_bc.cell(row=row_curr, column=y_idx)
                if pd.notna(val) and val is not None:
                    try:
                        c.value = float(val)
                        c.number_format = "#,##0"
                    except (ValueError, TypeError):
                        c.value = str(val)
                c.alignment = Alignment(horizontal="right")
                c.font = REG_FONT
                c.border = THIN_BORDER

            for c_idx in range(1, 5):
                ws_bc.cell(row=row_curr, column=c_idx).font = REG_FONT
                ws_bc.cell(row=row_curr, column=c_idx).border = THIN_BORDER
            row_curr += 1

    max_col_bc = len(headers_bc)
    max_row_bc = max(row_curr - 1, 4)
    ws_bc.auto_filter.ref = f"A4:{get_column_letter(max_col_bc)}{max_row_bc}"
    ws_bc.freeze_panes = "E5"

    ws_bc.column_dimensions["A"].width = 12
    ws_bc.column_dimensions["B"].width = 32
    ws_bc.column_dimensions["C"].width = 24
    ws_bc.column_dimensions["D"].width = 46
    for y_idx in range(5, max_col_bc + 1):
        ws_bc.column_dimensions[get_column_letter(y_idx)].width = 18

    # -----------------------------------------------------------------
    # Sheet 2: Ty_So_Tai_Chinh
    # -----------------------------------------------------------------
    if "Ty_So_Tai_Chinh" in wb.sheetnames:
        ws_ts = wb["Ty_So_Tai_Chinh"]
        if ws_ts.views.sheetView:
            ws_ts.views.sheetView[0].showGridLines = True
    else:
        ws_ts = wb.create_sheet("Ty_So_Tai_Chinh")

    ws_ts["A1"] = "CÁC TỶ SỐ TÀI CHÍNH PHÂN TÍCH (CHUẨN WIDATA)"
    ws_ts["A1"].font = TITLE_FONT

    ws_ts["A2"] = "Lọc theo Mã CK:"
    ws_ts["A2"].font = BOLD_FONT
    ws_ts["A2"].alignment = Alignment(horizontal="right")

    ws_ts["B2"] = "TẤT CẢ"
    ws_ts["B2"].font = BOLD_FONT
    ws_ts["B2"].alignment = Alignment(horizontal="center")
    ws_ts["B2"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws_ts.add_data_validation(dv)
    dv.add(ws_ts["B2"])

    headers_ts = ["Mã CK", "Phân nhóm tỷ số", "Mã chỉ số", "Tên chỉ số tài chính", "Công thức tính toán"] + [str(y) for y in years]
    for col_idx, h in enumerate(headers_ts, 1):
        cell = ws_ts.cell(row=4, column=col_idx, value=h)
        cell.font = WHITE_BOLD
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Xuất các chỉ số WiData (toàn bộ 75 chỉ số hoặc theo danh sách được chọn)
    active_ratios = list(WIDATA_RATIOS.keys())
    if fin_codebook:
        selected_ratios = [item.get("Biến") for item in fin_codebook if item.get("Phân loại") == "Tỷ số tài chính WiData"]
        if selected_ratios and len(selected_ratios) < len(WIDATA_RATIOS):
            active_ratios = [r for r in WIDATA_RATIOS if r in selected_ratios]


    row_ts = 5
    for t in tickers:
        df_t = pivot[pivot["ticker"] == t]
        for rcode in active_ratios:
            meta = WIDATA_RATIOS.get(rcode, {
                "name": rcode,
                "group": "Tỷ số tài chính",
                "formula": "Tính toán từ BCTC",
                "fmt": "0.00%",
            })
            ws_ts.cell(row=row_ts, column=1, value=t).alignment = Alignment(horizontal="center")
            ws_ts.cell(row=row_ts, column=2, value=meta["group"]).alignment = Alignment(horizontal="left")
            ws_ts.cell(row=row_ts, column=3, value=rcode).alignment = Alignment(horizontal="center")
            ws_ts.cell(row=row_ts, column=4, value=meta["name"]).alignment = Alignment(horizontal="left")
            ws_ts.cell(row=row_ts, column=5, value=meta["formula"]).alignment = Alignment(horizontal="left")

            for y_idx, y in enumerate(years, 6):
                row_match = df_t[df_t["year"] == y]
                val = row_match[rcode].values[0] if len(row_match) > 0 and rcode in row_match else None
                c = ws_ts.cell(row=row_ts, column=y_idx)
                if pd.notna(val) and val is not None:
                    c.value = float(val)
                    c.number_format = meta.get("fmt", "0.00%")
                c.alignment = Alignment(horizontal="right")
                c.font = REG_FONT
                c.border = THIN_BORDER

            for c_idx in range(1, 6):
                ws_ts.cell(row=row_ts, column=c_idx).font = REG_FONT
                ws_ts.cell(row=row_ts, column=c_idx).border = THIN_BORDER
            row_ts += 1

    max_col_ts = len(headers_ts)
    max_row_ts = max(row_ts - 1, 4)
    ws_ts.auto_filter.ref = f"A4:{get_column_letter(max_col_ts)}{max_row_ts}"
    ws_ts.freeze_panes = "F5"

    ws_ts.column_dimensions["A"].width = 12
    ws_ts.column_dimensions["B"].width = 26
    ws_ts.column_dimensions["C"].width = 24
    ws_ts.column_dimensions["D"].width = 46
    ws_ts.column_dimensions["E"].width = 46
    for y_idx in range(6, max_col_ts + 1):
        ws_ts.column_dimensions[get_column_letter(y_idx)].width = 18

    # -----------------------------------------------------------------
    # Sheet 3: Panel_Data_Goc
    # -----------------------------------------------------------------
    if "Panel_Data_Goc" in wb.sheetnames:
        ws_pnl = wb["Panel_Data_Goc"]
        if ws_pnl.views.sheetView:
            ws_pnl.views.sheetView[0].showGridLines = True
    else:
        ws_pnl = wb.create_sheet("Panel_Data_Goc")

    pnl_cols = list(pivot.columns)
    for c_idx, col_name in enumerate(pnl_cols, 1):
        cell = ws_pnl.cell(row=1, column=c_idx, value=col_name)
        cell.font = WHITE_BOLD
        cell.fill = TEAL_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for r_idx, (_, r) in enumerate(pivot.iterrows(), 2):
        for c_idx, col_name in enumerate(pnl_cols, 1):
            val = r[col_name]
            cell = ws_pnl.cell(row=r_idx, column=c_idx)
            if pd.notna(val) and val is not None:
                if isinstance(val, (int, float)):
                    cell.value = float(val)
                    if col_name == "year":
                        cell.number_format = "0"
                    elif col_name in ("ticker",):
                        pass
                    else:
                        cell.number_format = "#,##0.00" if abs(float(val)) < 100 else "#,##0"
                else:
                    cell.value = str(val)
            cell.font = REG_FONT
            cell.border = THIN_BORDER

    ws_pnl.auto_filter.ref = f"A1:{get_column_letter(len(pnl_cols))}{len(pivot) + 1}"
    ws_pnl.freeze_panes = "C2"
    ws_pnl.column_dimensions["A"].width = 14
    ws_pnl.column_dimensions["B"].width = 12
    for c_idx in range(3, min(len(pnl_cols) + 1, 60)):
        ws_pnl.column_dimensions[get_column_letter(c_idx)].width = 18

    # -----------------------------------------------------------------
    # Sheet 4: Codebook
    # -----------------------------------------------------------------
    if "Codebook" in wb.sheetnames:
        ws_cb = wb["Codebook"]
        if ws_cb.views.sheetView:
            ws_cb.views.sheetView[0].showGridLines = True
    else:
        ws_cb = wb.create_sheet("Codebook")

    cb_headers = ["Biến", "Tên chỉ tiêu", "Phân loại / Nhóm", "Phân loại", "Công thức / Nguồn"]
    for c_idx, h in enumerate(cb_headers, 1):
        cell = ws_cb.cell(row=1, column=c_idx, value=h)
        cell.font = WHITE_BOLD
        cell.fill = TEAL_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Xay dung codebook bao gom ca chi tieu BCTC va toan bo chi so WiData
    cb_full = list(fin_codebook)
    cb_vars = {item.get("Biến") for item in cb_full}

    for rk, rinfo in WIDATA_RATIOS.items():
        if rk not in cb_vars:
            cb_full.append({
                "Biến": rk,
                "Tên chỉ tiêu": rinfo["name"],
                "Phân loại / Nhóm": rinfo["group"],
                "Phân loại": "Tỷ số tài chính WiData",
                "Công thức / Nguồn": rinfo["formula"],
            })

    for r_idx, item in enumerate(cb_full, 2):
        for c_idx, h in enumerate(cb_headers, 1):
            val = item.get(h, "")
            cell = ws_cb.cell(row=r_idx, column=c_idx, value=val)
            cell.font = REG_FONT
            cell.border = THIN_BORDER
            if c_idx in (1, 3, 4):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    ws_cb.auto_filter.ref = f"A1:E{len(cb_full) + 1}"
    ws_cb.freeze_panes = "A2"
    ws_cb.column_dimensions["A"].width = 26
    ws_cb.column_dimensions["B"].width = 46
    ws_cb.column_dimensions["C"].width = 32
    ws_cb.column_dimensions["D"].width = 24
    ws_cb.column_dimensions["E"].width = 46

    # -----------------------------------------------------------------
    # Sheet 5: Huong_Dan_VBA
    # -----------------------------------------------------------------
    if "Huong_Dan_VBA" in wb.sheetnames:
        ws_hb = wb["Huong_Dan_VBA"]
        if ws_hb.views.sheetView:
            ws_hb.views.sheetView[0].showGridLines = True
    else:
        ws_hb = wb.create_sheet("Huong_Dan_VBA")

    ws_hb["A1"] = "HƯỚNG DẪN SỬ DỤNG BỘ LỌC VÀ MACRO VBA TRONG EXCEL"
    ws_hb["A1"].font = TITLE_FONT

    instructions = [
        ("1. Tổng quan các Tab trong bảng tính:", [
            "- Bao_Cao_Tai_Chinh: Trình bày toàn bộ 700+ chỉ tiêu BCTC dạng ngang, phân chia theo 13 nhóm chuẩn mực kế toán Việt Nam.",
            "- Ty_So_Tai_Chinh: Hệ thống chỉ số tài chính toàn diện theo chuẩn WiData (sinh lời, đòn bẩy, CTCK, YoY tăng trưởng).",
            "- Panel_Data_Goc: Bảng dữ liệu dạng bảng dài/rộng (Panel Data) phù hợp để chạy hồi quy định lượng trên Stata/R/Python.",
            "- Codebook: Từ điển định nghĩa chi tiết từng biến và nguồn gốc chỉ tiêu.",
        ]),
        ("2. Cách sử dụng bộ lọc mã chứng khoán (VBA Filter):", [
            "- Lưu ý bảo mật Office: Nếu Excel hiện thanh cảnh báo vàng, bấm 'Enable Content' (Kích hoạt nội dung).",
            "- Nếu Excel hiện thanh cảnh báo đỏ (Security Risk / Blocked): Đóng file -> Chuột phải vào file .xlsm trong thư mục tải về -> Chọn Properties -> Tích chọn 'Unblock' (Bỏ chặn) ở góc dưới tab General -> Bấm OK rồi mở lại file.",
            "- Cách 1 (Tự động): Chọn Mã CK tại ô B2 ở sheet 'Bao_Cao_Tai_Chinh' hoặc 'Ty_So_Tai_Chinh'. Bảng sẽ tự động lọc ngay lập tức.",
            "- Cách 2 (Nút bấm): Bấm nút [ Lọc Mã CK ], [ Hiện Tất Cả ], [ Đồng Bộ 2 Sheet ] được gắn sẵn trực tiếp ở dòng 2.",
            "- Cách 3 (Lọc gốc Excel không cần Macro): Bấm vào mũi tên AutoFilter trực tiếp tại ô A4 (cột 'Mã CK') để chọn bất kỳ doanh nghiệp nào.",
        ]),
        ("3. Các Macro có sẵn trong Module 'ModFilter':", [
            "- LocTheoMaCK: Lọc dữ liệu sheet hiện tại theo mã tại ô B2.",
            "- HienThiTatCa: Hủy lọc và hiển thị tất cả các công ty.",
            "- DongBoLocTatCaSheet: Lọc đồng thời cả 2 sheet Báo cáo và Tỷ số về cùng 1 mã CK được chọn.",
        ]),
    ]

    r_idx = 3
    for title, lines in instructions:
        ws_hb.cell(row=r_idx, column=1, value=title).font = SECTION_FONT
        r_idx += 1
        for line in lines:
            c = ws_hb.cell(row=r_idx, column=1, value="   " + line)
            c.font = REG_FONT
            r_idx += 1
        r_idx += 1

    ws_hb.column_dimensions["A"].width = 100

    return wb


def export_financial_workbooks(
    all_data: pd.DataFrame,
    pivot: pd.DataFrame,
    ratio_cols: Dict[str, str],
    fin_codebook: List[Dict[str, Any]],
    export_xlsx: Path,
    export_xlsm: Optional[Path] = None,
    template_xlsm: Optional[Path] = None,
) -> Dict[str, Path]:
    """Xuat song song file .xlsx chuan va file .xlsm co Macro VBA."""
    results: Dict[str, Path] = {}

    # 1. Tao file .xlsx chuan
    wb_xlsx = openpyxl.Workbook()
    populate_financial_sheets(wb_xlsx, all_data, pivot, ratio_cols, fin_codebook)
    if "Sheet" in wb_xlsx.sheetnames:
        del wb_xlsx["Sheet"]
    wb_xlsx.save(export_xlsx)
    wb_xlsx.close()
    results["xlsx"] = export_xlsx
    logger.success(f"Đã xuất file Excel chuẩn: {export_xlsx.name}")

    # 2. Tao file .xlsm co Macro VBA tu template
    if export_xlsm:
        if template_xlsm is None:
            template_xlsm = Path(__file__).resolve().parent.parent / "templates" / "vba_template.xlsm"

        if template_xlsm.exists():
            try:
                wb_xlsm = openpyxl.load_workbook(template_xlsm, keep_vba=True)
                populate_financial_sheets(wb_xlsm, all_data, pivot, ratio_cols, fin_codebook)
                wb_xlsm.save(export_xlsm)
                wb_xlsm.close()
                results["xlsm"] = export_xlsm
                logger.success(f"Đã xuất file Excel Macro VBA: {export_xlsm.name}")
            except Exception as e:
                logger.warning(f"Không thể xuất file .xlsm từ template: {e}")
        else:
            logger.warning(f"Không tìm thấy template VBA tại: {template_xlsm}")

    return results
