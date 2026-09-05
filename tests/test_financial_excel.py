# -*- coding: utf-8 -*-
"""
Unit tests for financial Excel export with transposed statements and VBA macros.
"""

from pathlib import Path
import openpyxl
import pandas as pd
import pytest

from arminer.export.financial_excel import export_financial_workbooks, populate_financial_sheets


def test_export_financial_workbooks(tmp_path: Path):
    tickers = ["VCB", "HPG"]
    years = [2022, 2023]
    records = []
    for t in tickers:
        for y in years:
            records.append({
                "ticker": t,
                "year": y,
                "statement": "balance_sheet",
                "item_code": "bs_tong_tai_san",
                "item_name": "TỔNG CỘNG TÀI SẢN",
                "value": 1500000.0,
            })
            records.append({
                "ticker": t,
                "year": y,
                "statement": "income_statement",
                "item_code": "is_doanh_thu_thuan",
                "item_name": "Doanh thu thuần",
                "value": 600000.0,
            })

    all_data = pd.DataFrame(records)
    pivot = all_data.pivot_table(index=["ticker", "year"], columns="item_code", values="value").reset_index()
    pivot["roa"] = 0.08
    pivot["roe"] = 0.18

    ratio_cols = {"roa": "ROA", "roe": "ROE"}
    fin_codebook = [
        {
            "Biến": "bs_tong_tai_san",
            "Tên chỉ tiêu": "TỔNG CỘNG TÀI SẢN",
            "Báo cáo / Nhóm": "Bảng cân đối kế toán",
            "Phân loại": "Chỉ tiêu kế toán",
            "Công thức / Nguồn": "vnfinancialdata",
        },
        {
            "Biến": "roa",
            "Tên chỉ tiêu": "ROA - Tỷ suất sinh lời trên tổng tài sản",
            "Báo cáo / Nhóm": "Tỷ số tài chính",
            "Phân loại": "Tỷ số tính toán",
            "Công thức / Nguồn": "Lợi nhuận sau thuế / Tổng tài sản",
        },
    ]

    export_xlsx = tmp_path / "financial_data.xlsx"
    export_xlsm = tmp_path / "financial_data.xlsm"

    results = export_financial_workbooks(
        all_data=all_data,
        pivot=pivot,
        ratio_cols=ratio_cols,
        fin_codebook=fin_codebook,
        export_xlsx=export_xlsx,
        export_xlsm=export_xlsm,
    )

    # 1. Check xlsx exists and has all 5 sheets
    assert export_xlsx.exists()
    wb_xlsx = openpyxl.load_workbook(export_xlsx)
    expected_sheets = ["Bao_Cao_Tai_Chinh", "Ty_So_Tai_Chinh", "Panel_Data_Goc", "Codebook", "Huong_Dan_VBA"]
    for s in expected_sheets:
        assert s in wb_xlsx.sheetnames

    # Check Bao_Cao_Tai_Chinh content
    ws_bc = wb_xlsx["Bao_Cao_Tai_Chinh"]
    assert ws_bc["A1"].value == "BÁO CÁO TÀI CHÍNH DOANH NGHIỆP"
    assert ws_bc["B2"].value == "TẤT CẢ"
    assert ws_bc.auto_filter.ref is not None
    # Row 4 headers: Mã CK, Phân nhóm báo cáo, Mã chỉ tiêu, Tên chỉ tiêu, 2022, 2023
    assert ws_bc.cell(4, 1).value == "Mã CK"
    assert ws_bc.cell(4, 2).value == "Phân nhóm báo cáo"
    assert ws_bc.cell(4, 3).value == "Mã chỉ tiêu"
    assert ws_bc.cell(4, 4).value == "Tên chỉ tiêu"

    # Check Ty_So_Tai_Chinh content
    ws_ts = wb_xlsx["Ty_So_Tai_Chinh"]
    assert "CÁC TỶ SỐ TÀI CHÍNH PHÂN TÍCH" in ws_ts["A1"].value
    assert ws_ts["B2"].value == "TẤT CẢ"
    assert ws_ts.cell(4, 1).value == "Mã CK"
    assert ws_ts.cell(4, 3).value == "Mã chỉ số"

    # Check xlsm exists if template was found
    if "xlsm" in results:
        assert export_xlsm.exists()
        wb_xlsm = openpyxl.load_workbook(export_xlsm, keep_vba=True)
        for s in expected_sheets:
            assert s in wb_xlsm.sheetnames
        wb_xlsm.close()

    wb_xlsx.close()


def test_full_702_indicators_guarantee(tmp_path: Path):
    import vnfinancialdata as vnf
    df_master = vnf.list_items(active_only=False)
    assert len(df_master) == 702

    # Tao data chi cho 2 chi tieu thuc te
    all_data = pd.DataFrame([
        {"ticker": "VCB", "year": 2023, "item_code": "bs_tong_tai_san", "item_name": "TỔNG CỘNG TÀI SẢN", "value": 1500000.0, "statement": "balance_sheet"},
        {"ticker": "VCB", "year": 2023, "item_code": "is_doanh_thu_thuan", "item_name": "Doanh thu thuần", "value": 600000.0, "statement": "income_statement"},
    ])
    pivot = all_data.pivot_table(index=["ticker", "year"], columns="item_code", values="value").reset_index()

    export_xlsx = tmp_path / "vcb_702.xlsx"
    # Goi xuat khong gioi han fin_codebook (de xuat toan bo 702 chi tieu mac dinh)
    results = export_financial_workbooks(
        all_data=all_data,
        pivot=pivot,
        ratio_cols={},
        fin_codebook=[],
        export_xlsx=export_xlsx,
    )

    assert export_xlsx.exists()
    wb = openpyxl.load_workbook(export_xlsx, data_only=True)
    ws_bc = wb["Bao_Cao_Tai_Chinh"]
    rows_bc = ws_bc.max_row - 4
    # Cam ket 100% dung 702 dong chi tieu cho ma VCB
    assert rows_bc == 702

    ws_ts = wb["Ty_So_Tai_Chinh"]
    rows_ts = ws_ts.max_row - 4
    # Cam ket 100% dung 75 dong chi so WiData
    assert rows_ts == 75

    # Kiem tra cac nhom ke toan deu co mat tren cot 2
    cats = {ws_bc.cell(r, 2).value for r in range(5, ws_bc.max_row + 1)}
    assert any("TÀI SẢN NGẮN HẠN" in str(c) for c in cats)
    assert any("VỐN CHỦ SỞ HỮU" in str(c) for c in cats)
    assert any("DOANH THU" in str(c) for c in cats)
    assert any("DÒNG TIỀN" in str(c) for c in cats)

    wb.close()

